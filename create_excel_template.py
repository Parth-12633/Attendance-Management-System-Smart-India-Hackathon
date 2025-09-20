from flask import Flask
from models import db, Student, User
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from app import app

# Get real student data from database
def get_student_data():
    with app.app_context():
        students = db.session.query(Student, User).join(User).filter(
            User.is_active == True
        ).order_by(Student.standard, Student.division, Student.roll_no).all()
        
        data = {
            'Roll No': [s.roll_no for s, _ in students],
            'Name': [u.name for _, u in students],
            'Class': [f"{s.standard}-{s.division}" for s, _ in students],
            'Status': [''] * len(students)
        }
        return data, students

# Create data dictionary
try:
    data, students = get_student_data()
except Exception as e:
    print("Could not fetch student data, using sample data instead")
    # Fallback to sample data
    data = {
        'Roll No': list(range(1, 11)),
        'Name': [
            'John Doe',
            'Jane Smith',
            'Alice Johnson',
            'Bob Wilson',
            'Charlie Brown',
            'Diana Prince',
            'Edward Stone',
            'Frank Miller',
            'Grace Lee',
            'Henry Ford'
        ],
        'Class': ['10-A'] * 10,
        'Status': [''] * 10
}

# Create DataFrame
df = pd.DataFrame(data)

# Save to Excel
excel_file = 'attendance_template.xlsx'
df.to_excel(excel_file, index=False)

# Load workbook for adding data validation
wb = load_workbook(excel_file)
ws = wb.active

# Create data validation for Status column
dv = DataValidation(
    type="list",
    formula1='"present,absent,late"',
    allow_blank=True
)

# Add the validation to the worksheet
ws.add_data_validation(dv)

# Apply validation to Status column (excluding header)
dv.add(f'D2:D{len(data["Roll No"]) + 1}')

# Adjust column widths
ws.column_dimensions['A'].width = 10  # Roll No
ws.column_dimensions['B'].width = 25  # Name
ws.column_dimensions['C'].width = 15  # Class
ws.column_dimensions['D'].width = 15  # Status

# Make header row bold
for cell in ws[1]:
    cell.font = cell.font.copy(bold=True)

# Save changes
wb.save(excel_file)

print(f"\nCreated attendance template successfully!")
print(f"File saved as: {excel_file}")
print("\nTemplate includes:")
if 'students' in locals():
    print(f"- {len(students)} actual students from database")
    # Print class-wise breakdown
    class_counts = {}
    for student, _ in students:
        class_key = f"{student.standard}-{student.division}"
        class_counts[class_key] = class_counts.get(class_key, 0) + 1
    
    print("\nClass-wise student count:")
    for class_name, count in sorted(class_counts.items()):
        print(f"Class {class_name}: {count} students")
else:
    print("- Sample data (10 students)")

print("\nInstructions:")
print("1. Open the Excel file")
print("2. Click on any cell in the Status column")
print("3. Use the dropdown arrow to select: present, absent, or late")
print("4. Fill in the status for each student")
print("5. Save the file when done\n")